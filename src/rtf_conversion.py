import re
import openpyxl
import datetime
import streamlit as st

def convert_graph(string_data, project_name):
    if 'TAKE #' in string_data:
        takes_list = string_data.split('TAKE #')
        num_takes = len(takes_list)
    else:
        num_takes = len([x for x in re.findall(re.escape('.')+"(.*?)"+re.escape('/'), string_data) if x != '' and ' ' not in x])
        takes_list = ['']
        for i in range(num_takes):
            take = string_data[string_data.find(f'.{i+1}/'):string_data.find(f'.{i+2}/')]
            takes_list.append(take)

    characters_in_takes_list = []
    st.session_state['num_takes'] = num_takes

    for take in takes_list:
        characters = get_characters_in_take(take)
        characters_in_takes_list.append(characters)
    
    st.session_state['characters_in_takes_list'] = characters_in_takes_list

    template = openpyxl.load_workbook('data/template.xlsx')

    for i in range(num_takes//50 + 1):
        write_xlsx_page(i, characters_in_takes_list, project_name, template)
    
    delete_empty_sheets(template, template.sheetnames)
    template.save('grafic_output.xlsx')

    characters_in_script = get_characters_in_script(characters_in_takes_list)
    characters_total_takes = get_total_num_takes_of_characters(characters_in_script, characters_in_takes_list)

    return characters_total_takes

def get_characters_in_take(take):
    characters = []
    for line in take.split('\n'):
        if '\t' in line:
            if line.count('*') > 2: # Deal with multiple characters in line
                multiple_chars = re.findall(re.escape('*')+"(.*?)"+re.escape('*'), line.split('\t')[0])
                for character in multiple_chars:
                    if character.isupper() or character.islower():
                        characters.append(clean_char(character))
            elif '/' in line:
                multiple_chars = line.split('\t')[0].split('/')
                for character in multiple_chars:
                    if character.isupper() or character.islower():
                        characters.append(clean_char(character))
            else:
                character = clean_char(line.split('\t')[0])
                if character.isupper() or character.islower():
                    characters.append(clean_char(character))
        if 'original' in line.lower():
            characters.append('ORIGINAL')
    #characters = re.findall(re.escape('*')+"(.*?)"+re.escape('*'), take)
    return list(set(characters))

def write_xlsx_page(page, characters_in_takes_list, project_name, template):
    sheets = template.sheetnames
    sheet = template[sheets[page]]

    sheet['D3'] = project_name # Nom del projecte
    sheet['AW3'] = datetime.datetime.now().strftime("%d/%m/%Y") # Data d'avui en format dd/mm/yyyy

    characters = get_characters_in_page(page, characters_in_takes_list)
    characters_total_takes = get_total_num_takes_of_characters(characters, characters_in_takes_list)
    characters_partial_takes = get_partial_num_takes_of_characters(characters, characters_in_takes_list, page)

    write_characters(sheet, characters)
    write_total_takes(sheet, characters, characters_total_takes)
    write_partial_takes(sheet, characters, characters_partial_takes)
    write_participation(sheet, characters, characters_in_takes_list, page)

    template.save('grafic_output.xlsx')

def get_characters_in_page(page, characters_in_takes_list):
    characters_in_page = []
    for i in range(50*page + 1, min(50*page + 50 + 1, len(characters_in_takes_list))):
        for character in characters_in_takes_list[i]:
            characters_in_page.append(character)
    return list(set(characters_in_page))

def get_characters_in_script(characters_in_takes_list):
    characters_in_script = []
    for i in range(1, len(characters_in_takes_list)):
        for character in characters_in_takes_list[i]:
            characters_in_script.append(character)
    return list(set(characters_in_script))

def get_total_num_takes_of_characters(characters, characters_in_takes_list):
    character_total_dict = {}
    for character in characters:
        for take in characters_in_takes_list:
            try:
                character_total_dict[character] += take.count(character)
            except:
                character_total_dict[character] = 0
    return character_total_dict

def get_partial_num_takes_of_characters(characters, characters_in_takes_list, page):
    character_partial_dict = {}
    for character in characters:
        for take in characters_in_takes_list[50*page : min(50*page + 50 + 1, len(characters_in_takes_list))]:
            try:
                character_partial_dict[character] += take.count(character)
            except:
                character_partial_dict[character] = 0
    return character_partial_dict

def write_characters(sheet, characters):
    for index, character in enumerate(characters):
        cell = f'D{8+index}'
        sheet[cell] = character

def write_total_takes(sheet, characters, characters_total_takes):
    for index, character in enumerate(characters):
        cell = f'A{8+index}'
        sheet[cell] = characters_total_takes[character]

def write_partial_takes(sheet, characters, characters_partial_takes):
    for index, character in enumerate(characters):
        cell = f'B{8+index}'
        sheet[cell] = characters_partial_takes[character]

def write_actors(sheet, characters, characters_actors_dict):
    for index, character in enumerate(characters):
        cell = f'C{8+index}'
        try:
            if characters_actors_dict[character] != '(buit)':
                sheet[cell] = characters_actors_dict[character]
        except:
            pass

def write_participation(sheet, characters, characters_in_takes_list, page):
    cell_columns = ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                'BA', 'BB']
    for index, character in enumerate(characters):
        row = 8+index
        for jndex, take in enumerate(characters_in_takes_list[50*page + 1 : min(50*page + 50 + 1, len(characters_in_takes_list))]):
            column = cell_columns[jndex]
            cell = f'{column}{row}'
            if character in take:
                if character.lower() != 'original':
                    border = openpyxl.styles.borders.Border(
                        diagonal= openpyxl.styles.borders.Side(style='thin'),
                        left=sheet[cell].border.left,
                        right=sheet[cell].border.right,
                        top=sheet[cell].border.top,
                        bottom=sheet[cell].border.bottom,
                        diagonalUp=True)
                    sheet[cell].border = border
                else:
                    sheet[cell] = 'O'

def delete_empty_sheets(workbook, sheets):
    for sheet in sheets:
        if workbook[sheet]['AW3'].value == None:
            workbook.remove_sheet(workbook[sheet])

def clean_char(character):
    return str(character).replace(':','').replace('*','').strip().upper()

def update_files(characters_total_takes):
    '''
    Add actors names to files
    '''
    characters_actors_dict = {}
    characters_in_takes_list = st.session_state['characters_in_takes_list']
    num_takes = st.session_state['num_takes']

    for character in list(characters_total_takes.keys()):
        characters_actors_dict[character] = st.session_state[character]
    st.session_state['characters_actors_dict'] = characters_actors_dict

    xlsx_file = openpyxl.load_workbook('grafic_output.xlsx')

    for i in range(num_takes//50 + 1):
        update_xlsx_page(i, characters_actors_dict, xlsx_file, characters_in_takes_list)
    
    update_summary(characters_total_takes)

def update_xlsx_page(page, characters_actors_dict, xlsx_file, characters_in_takes_list):
    sheets = xlsx_file.sheetnames
    sheet = xlsx_file[sheets[page]]

    characters = get_characters_in_page(page, characters_in_takes_list)

    write_actors(sheet, characters, characters_actors_dict)

    xlsx_file.save('grafic_output_updated.xlsx')

def convert_summary(characters_total_takes):
    project_name = st.session_state['project_name']
    characters_in_takes_list = st.session_state['characters_in_takes_list']

    with open('summary.txt', 'w') as f:
        f.write('================================\n')
        f.write(f'{project_name}\t{datetime.datetime.now().strftime("%d/%m/%Y")}\n')
        f.write('================================\n\n\n')
        f.write('/// RECOMPTE DE TAKES PER PERSONATGE\n\n')

        for character in list(characters_total_takes.keys()):
            takes_with_character_list = []
            f.write(f'{character} ({characters_total_takes[character]})\n')
            for index, take in enumerate(characters_in_takes_list):
                if character in take:
                    takes_with_character_list.append(str(index))
            takes_summary = ' | '.join(takes_with_character_list)
            f.write(f'{takes_summary}\n\n')

def update_summary(characters_total_takes):
    characters_in_takes_list = st.session_state['characters_in_takes_list']
    characters_actors_dict = st.session_state['characters_actors_dict']
    actors_takes_dict = {}
    
    old_summary = open("summary.txt", "r")
    old_summary_string = old_summary.read()

    with open('summary_updated.txt', 'w') as f:
        f.write(old_summary_string)
        f.write("/// RECOMPTE DE TAKES PER ACTORS DE DOBLATGE (pot ser que una mateixa take apareixi repetida si l'actor interpreta més d'un dels personatges que hi apareixen)\n\n")

        for character in list(characters_total_takes.keys()):
            takes_with_actor_list = []

            actor = characters_actors_dict[character]
            if actor not in actors_takes_dict:
                actors_takes_dict[actor] = []
            if actor != '' and actor != '(buit)':
                try: 
                    for index, take in enumerate(characters_in_takes_list):
                        if character in take:
                            takes_with_actor_list.append(str(index))
                    actors_takes_dict[actor].extend(takes_with_actor_list)
                except:
                    pass
        
        for actor in list(actors_takes_dict.keys()):
            if actor != '' and actor != '(buit)':
                f.write(f'{actor} ({len(actors_takes_dict[actor])})\n')
                takes_summary = ' | '.join(sorted(actors_takes_dict[actor], key=lambda x: float(x)))
                f.write(f'{takes_summary}\n\n')